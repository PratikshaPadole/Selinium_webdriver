import openpyxl
file="C:\\selenium_Practice\\test1.xlsx"
workbook = openpyxl.load_workbook(file)
sheet=workbook.active
sheet.cell(1,1).value=12
sheet.cell(1,2).value="pooja"
sheet.cell(1,3).value="shirur"
sheet.cell(2,1).value=13
sheet.cell(2,2).value="usha"
sheet.cell(2,3).value="pune"
sheet.cell(3,1).value=14
sheet.cell(3,2).value="niki"
sheet.cell(3,3).value="nandurbar"
sheet.cell(4,1).value=13
sheet.cell(4,2).value="usha"
sheet.cell(4,3).value="pune"

workbook.save(file)