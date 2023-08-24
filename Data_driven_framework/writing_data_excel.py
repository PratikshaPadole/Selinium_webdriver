import openpyxl
file="C:\\selenium_Practice\\sheet1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet = workbook.active
sheet.cell(1,1).value=123
sheet.cell(1,2).value="prati"
sheet.cell(1,3).value="automation engineer"

sheet.cell(2,1).value=143
sheet.cell(2,2).value="jan"
sheet.cell(2,3).value="lead engineer"

sheet.cell(3,1).value=456
sheet.cell(3,2).value="usha"
sheet.cell(3,3).value="full stack"
workbook.save(file)